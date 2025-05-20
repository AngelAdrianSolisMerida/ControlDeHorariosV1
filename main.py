import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
from datetime import datetime, timedelta, date
from tkcalendar import Calendar
import os
import openpyxl
import sys
import traceback

class HorariosExamenesApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Sistema de Horarios UNACH - Contaduría")
        self.root.geometry("1100x750")
        
        # Configuración de estilo
        self.style = ttk.Style()
        self.style.configure('TFrame', background='#f5f5f5')
        self.style.configure('TLabel', background='#f5f5f5', font=('Arial', 10))
        self.style.configure('TButton', font=('Arial', 10), background='#005f6a')
        self.style.configure('Header.TLabel', font=('Arial', 12, 'bold'), foreground='#005f6a')
        
        self.archivo_excel = None
        self.datos_procesados = None
        self.horarios_generados = None
        
        self.create_widgets()
    
    def create_widgets(self):
        # Frame principal
        main_frame = ttk.Frame(self.root, padding=(15, 15, 15, 15))
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(header_frame, 
                 text="UNIVERSIDAD AUTÓNOMA DE CHIAPAS\nSistema de Control de Horarios de Exámenes",
                 style='Header.TLabel', justify=tk.CENTER).pack(fill=tk.X)
        
        # Contenedor principal
        container = ttk.Frame(main_frame)
        container.pack(fill=tk.BOTH, expand=True)
        
        # Panel izquierdo (configuración)
        left_panel = ttk.LabelFrame(container, text=" Configuración ", padding=15)
        left_panel.pack(side=tk.LEFT, fill=tk.Y, padx=5, pady=5)
        
        # Panel derecho (resultados)
        right_panel = ttk.LabelFrame(container, text=" Resultados ", padding=15)
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Sección de archivo
        file_frame = ttk.LabelFrame(left_panel, text="1. Cargar Archivo Excel", padding=10)
        file_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(file_frame, text="Seleccionar archivo", 
                  command=self.cargar_archivo).pack(fill=tk.X)
        self.file_label = ttk.Label(file_frame, text="Ningún archivo seleccionado", 
                                  wraplength=250, foreground='#666666')
        self.file_label.pack(fill=tk.X, pady=5)
        
        # Sección de parámetros
        params_frame = ttk.LabelFrame(left_panel, text="2. Configurar Parámetros", padding=10)
        params_frame.pack(fill=tk.X, pady=5)
        
        # Tipo de examen
        ttk.Label(params_frame, text="Tipo de examen:").grid(row=0, column=0, sticky=tk.W, pady=2)
        self.exam_type = ttk.Combobox(params_frame, values=["Ordinario", "Extraordinario"], 
                                    state="readonly", width=15)
        self.exam_type.current(0)
        self.exam_type.grid(row=0, column=1, sticky=tk.EW, pady=2)
        self.exam_type.bind("<<ComboboxSelected>>", self.actualizar_config_examen)
        
        # Fechas
        ttk.Label(params_frame, text="Fecha inicio:").grid(row=1, column=0, sticky=tk.W, pady=2)
        self.start_date_btn = ttk.Button(params_frame, text="Seleccionar", 
                                       command=self.seleccionar_fecha_inicio, width=15)
        self.start_date_btn.grid(row=1, column=1, sticky=tk.EW, pady=2)
        self.start_date_label = ttk.Label(params_frame, text="No seleccionada", foreground='#666666')
        self.start_date_label.grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(0,5))
        
        ttk.Label(params_frame, text="Fecha fin:").grid(row=3, column=0, sticky=tk.W, pady=2)
        self.end_date_btn = ttk.Button(params_frame, text="Seleccionar", 
                                     command=self.seleccionar_fecha_fin, width=15)
        self.end_date_btn.grid(row=3, column=1, sticky=tk.EW, pady=2)
        self.end_date_label = ttk.Label(params_frame, text="No seleccionada", foreground='#666666')
        self.end_date_label.grid(row=4, column=0, columnspan=2, sticky=tk.W, pady=(0,5))
        
        # Configuración de examen
        ttk.Label(params_frame, text="Horas por examen:").grid(row=5, column=0, sticky=tk.W, pady=2)
        self.horas_label = ttk.Label(params_frame, text="3 horas")
        self.horas_label.grid(row=5, column=1, sticky=tk.W, pady=2)
        
        ttk.Label(params_frame, text="Exámenes por día:").grid(row=6, column=0, sticky=tk.W, pady=2)
        self.examenes_dia_label = ttk.Label(params_frame, text="2 matutino / 2 vespertino")
        self.examenes_dia_label.grid(row=6, column=1, sticky=tk.W, pady=2)
        
        # Sección de acciones
        action_frame = ttk.Frame(left_panel)
        action_frame.pack(fill=tk.X, pady=10)
        
        self.generate_btn = ttk.Button(action_frame, text="Generar Horarios", 
                                     command=self.generar_horarios, state=tk.DISABLED)
        self.generate_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        
        self.export_btn = ttk.Button(action_frame, text="Exportar a Excel", 
                                   command=self.exportar_excel, state=tk.DISABLED)
        self.export_btn.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)
        
        # Treeview para resultados
        self.tree = ttk.Treeview(right_panel, columns=("fecha", "hora", "materia", "docente", "semestre", "grupo", "turno"), 
                                show="headings", height=22)
        
        # Configurar columnas
        columns = [
            ("fecha", "Fecha", 100),
            ("hora", "Hora", 80),
            ("materia", "Materia", 250),
            ("docente", "Docente", 200),
            ("semestre", "Semestre", 80),
            ("grupo", "Grupo", 60),
            ("turno", "Turno", 100)
        ]
        
        for col_id, col_text, col_width in columns:
            self.tree.heading(col_id, text=col_text)
            self.tree.column(col_id, width=col_width, anchor=tk.CENTER if col_id in ["fecha", "hora", "semestre", "grupo", "turno"] else tk.W)
        
        # Scrollbars
        y_scroll = ttk.Scrollbar(right_panel, orient=tk.VERTICAL, command=self.tree.yview)
        x_scroll = ttk.Scrollbar(right_panel, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        
        # Grid layout
        self.tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        
        right_panel.grid_rowconfigure(0, weight=1)
        right_panel.grid_columnconfigure(0, weight=1)
        
        # Inicializar
        self.fecha_inicio = None
        self.fecha_fin = None
    
    def actualizar_config_examen(self, event=None):
        tipo = self.exam_type.get()
        if tipo == "Ordinario":
            self.horas_label.config(text="3 horas")
            self.examenes_dia_label.config(text="2 matutino / 2 vespertino")
        elif tipo == "Extraordinario":
            self.horas_label.config(text="2 horas")
            self.examenes_dia_label.config(text="3 matutino / 3 vespertino")
    
    def seleccionar_fecha_inicio(self):
        def set_fecha():
            try:
                selected_date = cal.selection_get()
                if selected_date < date.today():
                    messagebox.showerror("Error", "No se pueden seleccionar fechas pasadas")
                    return
                
                self.fecha_inicio = selected_date
                self.start_date_label.config(text=selected_date.strftime("%d/%m/%Y"))
                top.destroy()
                
                if self.fecha_fin and self.fecha_fin < self.fecha_inicio:
                    self.fecha_fin = None
                    self.end_date_label.config(text="No seleccionada")
            except Exception as e:
                messagebox.showerror("Error", f"Error al seleccionar fecha: {str(e)}")
                top.destroy()
        
        top = tk.Toplevel(self.root)
        top.title("Seleccionar Fecha Inicio")
        cal = Calendar(top, selectmode='day', mindate=date.today(), locale='es_ES',
                      font=('Arial', 10), headersbackground='#005f6a')
        cal.pack(padx=10, pady=10)
        ttk.Button(top, text="Aceptar", command=set_fecha).pack(pady=5)
    
    def seleccionar_fecha_fin(self):
        if not self.fecha_inicio:
            messagebox.showerror("Error", "Primero seleccione la fecha de inicio")
            return
            
        def set_fecha():
            try:
                selected_date = cal.selection_get()
                if selected_date < self.fecha_inicio:
                    messagebox.showerror("Error", "La fecha fin debe ser posterior a la fecha inicio")
                    return
                
                self.fecha_fin = selected_date
                self.end_date_label.config(text=selected_date.strftime("%d/%m/%Y"))
                top.destroy()
            except Exception as e:
                messagebox.showerror("Error", f"Error al seleccionar fecha: {str(e)}")
                top.destroy()
        
        top = tk.Toplevel(self.root)
        top.title("Seleccionar Fecha Fin")
        cal = Calendar(top, selectmode='day', mindate=self.fecha_inicio, locale='es_ES',
                      font=('Arial', 10), headersbackground='#005f6a')
        cal.pack(padx=10, pady=10)
        ttk.Button(top, text="Aceptar", command=set_fecha).pack(pady=5)
    
    def cargar_archivo(self):
        filepath = filedialog.askopenfilename(
            title="Seleccionar archivo Excel UNACH",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
            initialdir=os.path.expanduser("~")
        )
        
        if not filepath:
            return
            
        self.archivo_excel = filepath
        self.file_label.config(text=os.path.basename(filepath))
        
        try:
            # Intento leer el archivo con openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            
            # Buscar la hoja que contiene los datos (ignorando mayúsculas/minúsculas)
            target_sheets = ['lc', 'horarios', 'datos', 'contaduría']
            sheet_name = None
            
            for name in wb.sheetnames:
                lower_name = name.lower()
                if any(target in lower_name for target in target_sheets):
                    sheet_name = name
                    break
            
            if not sheet_name:
                sheet_name = wb.sheetnames[0]  # Tomar la primera hoja si no se encuentra
            
            sheet = wb[sheet_name]
            
            # Leer datos buscando el encabezado correcto
            data_rows = []
            headers = []
            encontrado_encabezado = False
            
            # Mapeo de columnas requeridas con posibles nombres alternativos
            column_mapping = {
                'SEMESTRE': ['semestre', 'nivel', 'grado'],
                'MATERIA': ['materia', 'asignatura', 'clase', 'nombre'],
                'GRUPO': ['grupo', 'sección', 'seccion', 'clave'],
                'TURNO': ['turno', 'jornada', 'horario'],
                'DOCENTE': ['docente', 'profesor', 'maestro', 'catedrático']
            }
            
            # Buscar el encabezado en las primeras filas
            for row in sheet.iter_rows(values_only=True):
                if not any(row):  # Fila vacía
                    continue
                
                # Convertir a minúsculas y limpiar
                current_headers = [str(cell).strip().lower() if cell is not None else '' for cell in row]
                
                # Verificar si esta fila contiene los encabezados que necesitamos
                found_columns = {}
                for standard_col, alternatives in column_mapping.items():
                    for idx, header in enumerate(current_headers):
                        if any(alt in header for alt in alternatives):
                            found_columns[standard_col] = idx
                            break
                
                # Si encontramos al menos 3 de las columnas requeridas, asumimos que es el encabezado
                if len(found_columns) >= 3:
                    headers = [(standard_col, idx) for standard_col, idx in found_columns.items()]
                    encontrado_encabezado = True
                    break
            
            if not encontrado_encabezado:
                wb.close()
                messagebox.showerror(
                    "Error en el archivo",
                    "No se pudo identificar el encabezado de columnas.\n\n"
                    "El archivo debe contener al menos estas columnas:\n"
                    "- Semestre/Nivel\n"
                    "- Materia/Asignatura\n"
                    "- Grupo/Sección\n"
                    "- Turno/Jornada\n"
                    "- Docente/Profesor"
                )
                self.datos_procesados = None
                self.generate_btn.config(state=tk.DISABLED)
                return
            
            # Leer los datos
            data = []
            header_indices = {col: idx for col, idx in headers}
            
            for row in sheet.iter_rows(min_row=sheet.min_row + 1, values_only=True):
                if not any(row):  # Saltar filas vacías
                    continue
                
                row_data = {}
                valid_row = True
                
                for col, idx in headers:
                    if idx < len(row):
                        value = row[idx]
                        # Limpiar valores None o strings vacíos
                        if value is None or (isinstance(value, str) and not value.strip()):
                            valid_row = False
                            break
                        row_data[col] = value
                    else:
                        valid_row = False
                        break
                
                if valid_row and row_data:
                    data.append(row_data)
            
            wb.close()
            
            if not data:
                messagebox.showerror(
                    "Error en el archivo",
                    "No se encontraron datos válidos en el archivo.\n\n"
                    "Por favor verifique que el archivo contenga información "
                    "en las columnas requeridas."
                )
                self.datos_procesados = None
                self.generate_btn.config(state=tk.DISABLED)
                return
            
            # Crear DataFrame
            df = pd.DataFrame(data)
            
            # Limpieza adicional
            if 'DOCENTE' in df.columns:
                df['DOCENTE'] = df['DOCENTE'].astype(str).str.split(',').str[0].str.strip()
            
            # Asignar licenciatura
            df['LICENCIATURA'] = "LICENCIATURA EN CONTADURÍA"
            
            self.datos_procesados = df
            self.generate_btn.config(state=tk.NORMAL)
            
            messagebox.showinfo(
                "Éxito",
                f"Archivo cargado correctamente.\n\n"
                f"Registros encontrados: {len(df)}\n"
                f"Columnas identificadas: {', '.join(df.columns)}"
            )
            
        except PermissionError:
            messagebox.showerror(
                "Error de acceso",
                "No se pudo leer el archivo. Por favor:\n\n"
                "1. Cierre el archivo Excel si está abierto\n"
                "2. Verifique que tenga permisos de lectura\n"
                "3. Intente con una copia del archivo"
            )
            self.datos_procesados = None
            self.generate_btn.config(state=tk.DISABLED)
            
        except Exception as e:
            error_msg = f"Error inesperado al leer el archivo:\n\n{str(e)}"
            messagebox.showerror("Error crítico", error_msg)
            self.datos_procesados = None
            self.generate_btn.config(state=tk.DISABLED)
    
    def validar_fechas(self):
        if not self.fecha_inicio or not self.fecha_fin:
            messagebox.showerror("Error", "Seleccione ambas fechas (inicio y fin)")
            return False
        
        if self.fecha_fin < self.fecha_inicio:
            messagebox.showerror("Error", "La fecha fin debe ser posterior a la fecha inicio")
            return False
            
        return True
    
    
    def generar_horarios(self):
        from datetime import timedelta
        import traceback

        if not self.datos_procesados is None and self.validar_fechas():
            try:
                tipo_examen = self.exam_type.get()

                config = {
                    "Ordinario": {
                        "horas": 3,
                        "examenes_por_turno": 2,
                        "dias_recomendados": 5,
                        "horario_matutino": ["08:00", "11:00"],
                        "horario_vespertino": ["15:00", "18:00"]
                    },
                    "Extraordinario": {
                        "horas": 2,
                        "examenes_por_turno": 3,
                        "dias_recomendados": 3,
                        "horario_matutino": ["08:00", "10:00", "12:00"],
                        "horario_vespertino": ["15:00", "17:00", "19:00"]
                    }
                }[tipo_examen]

                dias_habiles = []
                current_date = self.fecha_inicio
                while current_date <= self.fecha_fin:
                    if current_date.weekday() < 5:
                        dias_habiles.append(current_date)
                    current_date += timedelta(days=1)

                if not dias_habiles:
                    messagebox.showerror("Error", "No hay días hábiles en el rango de fechas seleccionado")
                    return

                if len(dias_habiles) < config["dias_recomendados"]:
                    messagebox.showwarning("Advertencia", f"Se recomiendan al menos {config['dias_recomendados']} días hábiles para exámenes {tipo_examen}.Actualmente hay {len(dias_habiles)} días hábiles.")

                df = self.datos_procesados.copy()
                df = df.dropna(subset=['SEMESTRE', 'MATERIA', 'GRUPO', 'TURNO', 'DOCENTE'])
                df['SEMESTRE'] = df['SEMESTRE'].astype(str).str.strip()
                df['GRUPO'] = df['GRUPO'].astype(str).str.strip().str.upper()
                df['TURNO'] = df['TURNO'].astype(str).str.strip().str.upper()
                df['DOCENTE'] = df['DOCENTE'].astype(str).str.split(',').str[0].str.strip()
                df['ES_OPTATIVA'] = df['MATERIA'].str.contains('OPTATIVA', case=False, na=False)

                if df.empty:
                    messagebox.showerror("Error", "No hay datos válidos después de la limpieza")
                    return

                todos_semestres = sorted(df['SEMESTRE'].unique(), key=lambda x: int(x) if x.isdigit() else x)
                todos_grupos = sorted(df['GRUPO'].unique())
                materias_por_grupo = {}
                for semestre in todos_semestres:
                    materias_por_grupo[semestre] = {}
                    for grupo in todos_grupos:
                        materias_grupo = df[(df['SEMESTRE'] == semestre) & (df['GRUPO'] == grupo)]
                        if not materias_grupo.empty:
                            materias_por_grupo[semestre][grupo] = materias_grupo

                docentes_asignados = {}
                grupos_asignados = {}

                def docente_disponible(docente, fecha, hora):
                    return not (fecha in docentes_asignados and docente in docentes_asignados[fecha] and hora in docentes_asignados[fecha][docente])

                def grupo_disponible(grupo, fecha, hora, es_optativa):
                    if es_optativa:
                        return True
                    return not (fecha in grupos_asignados and grupo in grupos_asignados[fecha] and hora in grupos_asignados[fecha][grupo])

                horarios = []
                for semestre in todos_semestres:
                    for grupo, materias_grupo in materias_por_grupo.get(semestre, {}).items():
                        matutinas = materias_grupo[materias_grupo['TURNO'] == 'M']
                        vespertinas = materias_grupo[materias_grupo['TURNO'] == 'V']

                        for i in range(min(config["examenes_por_turno"], len(matutinas))):
                            materia = matutinas.iloc[i]
                            hora = config["horario_matutino"][i]
                            for dia in dias_habiles:
                                if docente_disponible(materia['DOCENTE'], dia, hora) and grupo_disponible(grupo, dia, hora, materia['ES_OPTATIVA']):
                                    docentes_asignados.setdefault(dia, {}).setdefault(materia['DOCENTE'], []).append(hora)
                                    if not materia['ES_OPTATIVA']:
                                        grupos_asignados.setdefault(dia, {}).setdefault(grupo, []).append(hora)
                                    horarios.append({
                                        "Fecha": dia,
                                        "Hora": hora,
                                        "Licenciatura": "LICENCIATURA EN CONTADURÍA",
                                        "Materia": materia['MATERIA'],
                                        "Docente": materia['DOCENTE'],
                                        "Semestre": materia['SEMESTRE'],
                                        "Grupo": materia['GRUPO'],
                                        "Turno": "Matutino",
                                        "EsOptativa": materia['ES_OPTATIVA']
                                    })
                                    break

                        for i in range(min(config["examenes_por_turno"], len(vespertinas))):
                            materia = vespertinas.iloc[i]
                            hora = config["horario_vespertino"][i]
                            for dia in dias_habiles:
                                if docente_disponible(materia['DOCENTE'], dia, hora) and grupo_disponible(grupo, dia, hora, materia['ES_OPTATIVA']):
                                    docentes_asignados.setdefault(dia, {}).setdefault(materia['DOCENTE'], []).append(hora)
                                    if not materia['ES_OPTATIVA']:
                                        grupos_asignados.setdefault(dia, {}).setdefault(grupo, []).append(hora)
                                    horarios.append({
                                        "Fecha": dia,
                                        "Hora": hora,
                                        "Licenciatura": "LICENCIATURA EN CONTADURÍA",
                                        "Materia": materia['MATERIA'],
                                        "Docente": materia['DOCENTE'],
                                        "Semestre": materia['SEMESTRE'],
                                        "Grupo": materia['GRUPO'],
                                        "Turno": "Vespertino",
                                        "EsOptativa": materia['ES_OPTATIVA']
                                    })
                                    break

                if not horarios:
                    messagebox.showwarning("Advertencia", "No se generaron horarios. Verifique los datos de entrada.")
                    return

                self.horarios_generados = pd.DataFrame(horarios)
                self.horarios_generados['HoraOrden'] = pd.to_datetime(self.horarios_generados['Hora'], format='%H:%M')
                self.horarios_generados = self.horarios_generados.sort_values(['Fecha', 'HoraOrden', 'Semestre', 'Grupo'])
                self.horarios_generados = self.horarios_generados.drop('HoraOrden', axis=1)
                self.mostrar_resultados()
                self.export_btn.config(state=tk.NORMAL)
                messagebox.showinfo("Horarios Generados", f"Total: {len(horarios)} exámenes programados.")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudieron generar los horarios:{str(e)}")

    def mostrar_resultados(self):
        # Limpiar treeview
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        if self.horarios_generados is None or self.horarios_generados.empty:
            return
        
        # Ordenar por fecha y hora
        df = self.horarios_generados.sort_values(["Fecha", "Hora"])
        
        # Mostrar en el treeview con formato
        for _, row in df.iterrows():
            self.tree.insert("", tk.END, values=(
                row["Fecha"].strftime("%d/%m/%Y") if hasattr(row["Fecha"], 'strftime') else row["Fecha"],
                row["Hora"],
                row["Materia"],
                row["Docente"],
                row["Semestre"],
                row["Grupo"],
                row["Turno"]
            ))
    
    def exportar_excel(self):
        if self.horarios_generados is None or self.horarios_generados.empty:
            messagebox.showerror("Error", "No hay horarios generados para exportar")
            return
            
        default_name = f"Horarios_{self.exam_type.get()}_{date.today().strftime('%Y%m%d')}.xlsx"
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Guardar horarios como",
            initialfile=default_name
        )
        
        if not filepath:
            return
            
        try:
            # Crear un DataFrame ordenado
            df_export = self.horarios_generados.sort_values(["Fecha", "Hora"])
            
            # Convertir la columna Fecha a string si es datetime
            if pd.api.types.is_datetime64_any_dtype(df_export["Fecha"]):
                df_export["Fecha"] = df_export["Fecha"].dt.strftime("%d/%m/%Y")
            elif hasattr(df_export["Fecha"].iloc[0], 'strftime'):
                df_export["Fecha"] = df_export["Fecha"].apply(lambda x: x.strftime("%d/%m/%Y"))
            
            # Crear archivo Excel
            df_export.to_excel(filepath, index=False, sheet_name="Horarios")
            
            messagebox.showinfo(
                "Éxito",
                f"Archivo guardado correctamente:\n\n{filepath}\n\n"
                f"Total de registros: {len(df_export)}"
            )
            
        except PermissionError:
            messagebox.showerror(
                "Error de acceso",
                "No se pudo guardar el archivo. Por favor:\n\n"
                "1. Cierre el archivo Excel si está abierto\n"
                "2. Verifique que tenga permisos de escritura\n"
                "3. Intente con otra ubicación"
            )
        except Exception as e:
            messagebox.showerror(
                "Error al guardar",
                f"No se pudo guardar el archivo:\n\n{str(e)}\n\n"
                f"Tipo de error: {type(e).__name__}"
            )

if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = HorariosExamenesApp(root)
        root.mainloop()
    except Exception as e:
        messagebox.showerror(
            "Error crítico", 
            f"El programa no pudo iniciar:\n\n{str(e)}"
        )
        sys.exit(1)